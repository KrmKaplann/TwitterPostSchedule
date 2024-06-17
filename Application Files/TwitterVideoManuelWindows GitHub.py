# -*- coding: utf-8 -*-

import json
import os
import undetected_chromedriver as uc
import time
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import locale
import pyautogui
from datetime import datetime
import re

locale.setlocale(locale.LC_TIME, "tr_TR.UTF-8")

def Driver():
    options = uc.ChromeOptions()
    options.add_argument("--disable-popup-blocking")
    #options.add_argument("--incognito")
    #options.add_argument("--headless")  # Run the browser in headless mode
    options.add_argument("--start-maximized")  # Start the browser maximized
    driver = uc.Chrome(options=options)
    return driver

def Chrome_Tab(driver, link):
    driver.get(link)
    driver.maximize_window()

path = "<ExcelFilePath>"
# <ExcelFilePath>: Path to the Excel file
workbook = openpyxl.load_workbook(path)
DefaultPageSheet = workbook["<SheetName>"]
# <SheetName>: Name of the sheet in the Excel file

IslemYapilacakSosyalMedya = "Twitter"

StartColumn = 5
while True:
    Finder = DefaultPageSheet.cell(2, StartColumn).value
    if Finder == IslemYapilacakSosyalMedya:
        break
    StartColumn += 1

TotalAccountListGmail = []
DefaultRow = 4

while True:
    Start = DefaultPageSheet.cell(DefaultRow, StartColumn).value
    if Start is None:
        break
    if DefaultPageSheet.cell(DefaultRow, StartColumn + 1).value != "-":
        TotalAccountListGmail.append([Start, DefaultRow - 3])
    DefaultRow += 1

# Filter out entries with '-'
TotalAccountListGmail = [item for item in TotalAccountListGmail if item[0] != '-']

TotalAccountList = []
DefaultRow = 4

for account_info in TotalAccountListGmail:
    index = account_info[1]
    Start = DefaultPageSheet.cell(index + 3, 2).value
    TotalAccountList.append(Start)

print(TotalAccountList)

# Get the starting index from the user
baslangic_indeksi = int(input("Enter the last completed index for Twitter accounts (0 to {}): ".format(len(TotalAccountList)-1)))

TotalAccountList = TotalAccountList[baslangic_indeksi:]
TotalAccountListGmail = TotalAccountListGmail[baslangic_indeksi:]

path = "<TwitterExcelFilePath>"
# <TwitterExcelFilePath>: Path to the Twitter Excel file
workbook = openpyxl.load_workbook(path)
AllPagesWorksheet = workbook.sheetnames

for IndexAccount, OneItem in enumerate(TotalAccountListGmail, start=0):
    driver = Driver()
    link = "https://www.twitter.com"

    Chrome_Tab(driver, link)

    JsonFileName = DefaultPageSheet.cell(2, StartColumn).value + OneItem[0] + ".json"
    with open(f"<APIsPath>/{JsonFileName}", "r") as file:
        # <APIsPath>: Path to the APIs directory
        cookies = json.load(file)

    for cookie in cookies:
        driver.add_cookie(cookie)

    time.sleep(3)
    driver.refresh()

    for index, OnePage in enumerate(TotalAccountList, start=1):
        print(f"[ {index} ] {OnePage}")

    StartRow = 5
    PostDict = {}

    sekmeler = driver.window_handles
    ilk_sekme = sekmeler[0]
    driver.switch_to.window(ilk_sekme)

    time.sleep(2)
    if StartRow == 5:
        for i in range(3):
            time.sleep(1)
            pyautogui.hotkey('ctrl', '-')
    time.sleep(1)

    OneAccountName = TotalAccountList[IndexAccount]
    StartRow = 5
    AllPostList = []
    while True:
        SpecialWorkSheet = workbook[OneAccountName]
        PostName = SpecialWorkSheet["B" + str(StartRow)].value
        StartRow += 1
        if PostName is None:
            break
        AllPostList.append(PostName)

    print(AllPostList)
    AllPostDefault = AllPostList

    baslangic_indeksi = int(input(f"Enter the last completed index for {OneAccountName} (0 to {(len(AllPostList) - 1)}): "))

    AllPostList = AllPostList[baslangic_indeksi:]

    for StartRow, PostOne in enumerate(AllPostDefault, start=5):
        SpecialWorkSheet = workbook[OneAccountName]
        PostName = SpecialWorkSheet["B" + str(StartRow)].value
        Aciklama = SpecialWorkSheet["C" + str(StartRow)].value

        Tarih = SpecialWorkSheet["D" + str(StartRow)].value
        Tarih = datetime.strptime(Tarih, '%d.%m.%Y')
        FormatliTarih = Tarih.strftime("%d %B %A %Y")
        Gun = Tarih.strftime("%d")
        Ay = Tarih.strftime("%B")
        Saat = str(SpecialWorkSheet["E" + str(StartRow)].value)
        SaatSplitted = Saat.split(":")
        Hour = str(SaatSplitted[0])
        Minutes = str(SaatSplitted[1])
        HemenPaylas = SpecialWorkSheet["F" + str(StartRow)].value

        post_details = {
            "Aciklama": Aciklama,
            "Tarih": FormatliTarih,
            "Gün": Gun,
            "Ay": Ay,
            "Saat": Saat,
            "Hour": Hour,
            "Minutes": Minutes,
            "HemenPaylas": HemenPaylas
        }

        if OneAccountName in PostDict:
            PostDict[OneAccountName][PostName] = post_details
        else:
            PostDict[OneAccountName] = {PostName: post_details}

    for StartRow, PostOne in enumerate(AllPostList, start=5):
        wait = WebDriverWait(driver, 5)
        driver.implicitly_wait(180)

        LogoVisibility = wait.until(EC.visibility_of_element_located((By.XPATH, "<AvatarButton>")))
        # <AvatarButton>: XPath for avatar button
        cleaned_text = re.sub(r"/Başlık", "", str(PostDict[OneAccountName][PostOne]["Aciklama"]))
        driver.find_element(By.XPATH, "<TweetInput>").send_keys(cleaned_text)
        # <TweetInput>: XPath for tweet input
        FileAdress = "<TwitterVideoPath>/" + OneAccountName + "/" + PostOne + ".mp4"
        # <TwitterVideoPath>: Path to the Twitter videos

        driver.find_element(By.XPATH, "<FileInput>").send_keys(FileAdress)
        # <FileInput>: XPath for file input
        UploadWait = wait.until(EC.visibility_of_element_located((By.XPATH, "<UploadCompleteText>")))
        # <UploadCompleteText>: XPath for upload complete text

        time.sleep(10)
        driver.find_element(By.XPATH, "<ScheduleTweetButton>").click()
        # <ScheduleTweetButton>: XPath for schedule tweet button

        Month = str(PostDict[OneAccountName][PostOne]["Ay"])
        if "May" in Month:
            Month = "Mayıs"
        time.sleep(2)
        driver.find_element(By.XPATH, "<MonthSelect>").click()
        # <MonthSelect>: XPath for month select

        DateNo = str(PostDict[OneAccountName][PostOne]["Gün"])
        if "0" in DateNo[0]:
            DateNo = DateNo[1]

        for i in range(1):
            try:
                time.sleep(1)
                driver.find_element(By.XPATH, "<DaySelect>").click()
                # <DaySelect>: XPath for day select
            except:
                pass

        time.sleep(2)
        driver.find_element(By.XPATH, "<YearSelect>").click()
        # <YearSelect>: XPath for year select

        Saat = str(PostDict[OneAccountName][PostOne]["Hour"])
        driver.find_element(By.XPATH, "<HourSelect>").click()
        # <HourSelect>: XPath for hour select

        Minutes = str(PostDict[OneAccountName][PostOne]["Minutes"])
        time.sleep(2)
        driver.find_element(By.XPATH, "<MinuteSelect>").click()
        # <MinuteSelect>: XPath for minute select

        time.sleep(2)
        driver.find_element(By.XPATH, "<ScheduleConfirmButton>").click()
        # <ScheduleConfirmButton>: XPath for schedule confirm button

        time.sleep(2)
        driver.find_element(By.XPATH, "<TweetButton>").click()
        # <TweetButton>: XPath for tweet button

        time.sleep(5)

        driver.refresh()
        time.sleep(3)

    driver.quit()
print("Completed...")
